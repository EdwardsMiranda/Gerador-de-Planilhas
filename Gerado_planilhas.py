import openpyxl

planilha = openpyxl.Workbook()

print('Seja bem vindo ao criador de planilhas')
print('='*60)
print('Para começar, vamos criar uma página dentro de uma planilha.')

#1. Criar novas páginas(quantas a usuário quiser)
adicionar_nova_planilha = True
while adicionar_nova_planilha == True:
    planilha.create_sheet(input('Digite o nome da planilha  '))
    resposta_planilha = input('Deseja criar mais uma página? (s/n)  ')
    print('='*60)
    if resposta_planilha =='s':
        adicionar_nova_planilha = True
    else:
        adicionar_nova_planilha = False
        print(planilha.sheetnames)

#2. Escolher entre as páginas criadas anteriormente, qual página deve ser modificada
pagina_escolhida = planilha[input('Digite o nome da página que deseja manipular: ')]
print('='*60)

#3. Após escolher qual página será modificada, permitir que o usuário adicione um cabeçalho a esta
adicionar_nova_coluna = True
while adicionar_nova_coluna == True:
    pagina_escolhida.append([input('Digite uma coluna para o cabeçalho: ')])
    print('='*60)
    resposta_coluna = input('Deseja criar mais uma coluna? (s/n)  ')
    print('='*60)
    if resposta_coluna == 's':
        adicionar_nova_coluna = True
    else:
        adicionar_nova_coluna = False

#4. opção de adicionar dados a planilha ou simplesmente fechar o programa
adicionar_novos_dados = True
while adicionar_novos_dados == True:
    resposta_dados = input('Deseja adicionar dados a essa planilha? (s/n) ')
    print('='*60)
    if resposta_dados == 's':
        
        print('As páginas disponíveis são: ' + str(planilha.sheetnames))
        print('='*60)
        pagina_escolhida2 = planilha[input('Em qual página você deseja adicionar dados? ')]
        print('='*60)
        adicionar_novos_dados = True

        adicionar_nova_linha = True
        while adicionar_nova_linha ==True:
            pagina_escolhida2.append([input('Digite os dados a serem adicionado a uma nova linha, SEPARADOS POR VÍRGULA: ')])
            print('='*60)
            resposta_linha = input('Deseja adicionar nova linha? s/n  ')
            print('='*60)
            if resposta_linha == 's':
                adicionar_nova_linha = True
            else:
                adicionar_nova_linha = False
    else:
        adicionar_novos_dados = False
        planilha.save(input('Digite o nome da plan1ilha que deseja salvar: ') + '.xlsx')
        print('='*60)
planilha.close()
print('Planilha salva com sucesso!')


#9. Para finalizar, o programa deverá perguntar ao usuário qual é o nome da planilha a ser salva e,
#depois de salvar a planilha, enviar uma mensagem que o programa está sendo finalizado e depois
#finalizar a execução do programa.

